import os
import requests
import openpyxl
import json
import datetime
from bs4 import BeautifulSoup
from urllib.parse import urlencode
from sqlalchemy.orm import Session
from config.config import SHEDULE

from database import tables


def get_urls():
    urls = []
    url = SHEDULE.URL
    r = requests.get(url)
    soup = BeautifulSoup(r.text, 'html.parser')
    p = soup.findAll('div', class_='container')[2].findAll('a')
    for url in p:

        if url.text.lower().find('расписание') >= 0:
            urls.append(url.get('href'))
    return urls


def download_file(url, output_filename):
    base_url = SHEDULE.YANDEX_DOWNLOAD_URL

    # Получаем загрузочную ссылку
    final_url = base_url + urlencode(dict(public_key=url))
    response = requests.get(final_url)
    download_url = response.json()['href']

    # Загружаем файл и сохраняем его
    download_response = requests.get(download_url)
    with open(output_filename, 'wb') as f:
        f.write(download_response.content)
    return output_filename


def parse_excel(filename):
    workbook = openpyxl.load_workbook(filename)
    sheets = workbook.sheetnames

    worksheet = workbook[sheets[0]]
    rows = []
    row = 12

    global ID
    while (worksheet.cell(row=row, column=2).value is not None) and (
            worksheet.cell(row=row, column=4).value is not None):
        if worksheet.cell(row=row, column=6).value is not None:
            rows.append({
                'ID': ID,
                'date': str(worksheet.cell(row=row, column=2).value)[:10],
                'day_of_week': worksheet.cell(row=row, column=3).value,
                'time': str(worksheet.cell(row=row, column=4).value),
                'group': worksheet.cell(row=row, column=5).value,
                'subject': worksheet.cell(row=row, column=6).value,
                'teacher': str(worksheet.cell(row=row, column=7).value).replace("\n", ''),
                'department': worksheet.cell(row=row, column=8).value,
                'class_type': worksheet.cell(row=row, column=9).value,
                'auditory': worksheet.cell(row=row, column=10).value
            })
            ID += 1
        row += 1

    data = {}
    group_list = []

    for row in rows:
        group = row['group']
        if group not in data:
            data.update({f'{group}': []})
        data[f'{group}'].append(row)

    for group in data:
        group_list.append(group)
        arr = data[f'{group}']
        date = datetime.date.today().strftime('%d.%m.%Y')
        path = os.getcwd().replace('\\', '/')
        if not os.path.isdir(path + f"/{SHEDULE.DIRECTORY}/"):
            os.mkdir(path + f"/{SHEDULE.DIRECTORY}/")
        if not os.path.isdir(path + f"/{SHEDULE.DIRECTORY}/{date}"):
            os.mkdir(path + f"/{SHEDULE.DIRECTORY}/{date}")
        with open(f"./{SHEDULE.DIRECTORY}/{date}/{group}.json", 'w', encoding='utf-8') as file:
            file.write(json.dumps(arr, ensure_ascii=False))
    return group_list


def parse_shedule():
    urls = get_urls()
    group_list = []
    global ID
    ID = 1

    for url in urls:
        filename = download_file(url, 'shedule.xlsx')
        group_list.extend(parse_excel(filename))

    os.remove('shedule.xlsx')
    return group_list


def update_shedule(session: Session):
    path = f"{SHEDULE.DIRECTORY}/{datetime.date.today().strftime('%d.%m.%Y')}/"
    group_list = parse_shedule()
    for group_name in group_list:
        with open(path + group_name + '.json', encoding='utf-8') as f:
            group_shedule = json.load(f)
            group = session.query(tables.Group).filter(tables.Group.group_name == group_name).first()
            if group:
                group.group_shedule = group_shedule
                session.commit()
            else:
                group = tables.Group(
                    group_name=group_name,
                    group_shedule=group_shedule
                )
                session.add(group)
                session.commit()
